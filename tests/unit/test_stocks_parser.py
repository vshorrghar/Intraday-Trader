"""Unit tests for parsers.groww_stocks_parser."""

import logging

import openpyxl
import pytest

from parsers.groww_stocks_parser import (
    KNOWN_INVIT_ISINS,
    classify_holding,
    parse_stocks_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_stocks_xlsx(path, headers=None, rows=None):
    """Create a minimal Groww-style stocks XLSX at *path*.

    Headers go into row 11 (columns A-H).  Data rows start at row 12.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    if headers is None:
        headers = [
            "Stock Name", "ISIN", "Quantity", "Average buy price",
            "Buy value", "Closing price", "Closing value", "Unrealised P&L",
        ]

    for col_idx, hdr in enumerate(headers, start=1):  # A=1
        ws.cell(row=11, column=col_idx, value=hdr)

    if rows:
        for row_offset, row_data in enumerate(rows):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=12 + row_offset, column=col_idx, value=val)

    wb.save(path)


# ---------------------------------------------------------------------------
# classify_holding tests
# ---------------------------------------------------------------------------

class TestClassifyHolding:
    """Tests for classify_holding()."""

    def test_stock_ine_prefix(self):
        assert classify_holding("INE002A01018", "Reliance Industries", set()) == "stock"

    def test_etf_inf_prefix(self):
        assert classify_holding("INF204KB17I5", "Motilal Oswal Nasdaq 100", set()) == "etf"

    def test_etf_name_keyword_etf(self):
        assert classify_holding("INE123X01000", "Nippon India ETF Nifty 50", set()) == "etf"

    def test_etf_name_keyword_bees(self):
        assert classify_holding("INE123X01000", "Nippon India Nifty Bees", set()) == "etf"

    def test_etf_name_keyword_nasdaq(self):
        assert classify_holding("INE123X01000", "Motilal Oswal Nasdaq Fund", set()) == "etf"

    def test_etf_name_keyword_mafang(self):
        assert classify_holding("INE123X01000", "Mirae MAFANG Fund", set()) == "etf"

    def test_etf_name_keyword_mahktech(self):
        assert classify_holding("INE123X01000", "Mirae MAHKTECH Fund", set()) == "etf"

    def test_etf_name_keyword_silver(self):
        assert classify_holding("INE123X01000", "ICICI Silver Fund", set()) == "etf"

    def test_etf_name_case_insensitive(self):
        assert classify_holding("INE123X01000", "some etf fund", set()) == "etf"

    def test_invit_classification(self):
        invit_isins = {"INE183W23014"}
        assert classify_holding("INE183W23014", "Powergrid InvIT", invit_isins) == "invit"

    def test_invit_takes_priority_over_stock(self):
        """InvIT ISINs start with INE but should be classified as invit, not stock."""
        invit_isins = {"INE0Z8Z23013"}
        assert classify_holding("INE0Z8Z23013", "Brookfield REIT", invit_isins) == "invit"

    def test_returns_exactly_one_of_three(self):
        for result in [
            classify_holding("INE002A01018", "Reliance", set()),
            classify_holding("INF204KB17I5", "Motilal ETF", set()),
            classify_holding("INE183W23014", "Powergrid InvIT", {"INE183W23014"}),
        ]:
            assert result in {"stock", "etf", "invit"}


# ---------------------------------------------------------------------------
# parse_stocks_xlsx tests
# ---------------------------------------------------------------------------

class TestParseStocksXlsx:
    """Tests for parse_stocks_xlsx()."""

    def test_valid_parsing(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Reliance Industries", "INE002A01018", 10, 2500.0, 25000.0, 2600.0, 26000.0, 1000.0],
            ["HDFC Bank", "INE040A01034", 5, 1600.0, 8000.0, 1650.0, 8250.0, 250.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 2
        assert holdings[0].name == "Reliance Industries"
        assert holdings[0].isin == "INE002A01018"
        assert holdings[0].quantity == 10
        assert holdings[0].avg_buy_price == 2500.0
        assert holdings[0].buy_value == 25000.0
        assert holdings[0].groww_closing_price == 2600.0
        assert holdings[0].groww_closing_value == 26000.0
        assert holdings[0].unrealised_pnl == 1000.0
        assert holdings[0].holding_type == "stock"

    def test_etf_classification_in_parsing(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Motilal Oswal Nasdaq 100", "INF204KB17I5", 20, 100.0, 2000.0, 110.0, 2200.0, 200.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].holding_type == "etf"

    def test_invit_classification_in_parsing(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Powergrid InvIT", "INE183W23014", 100, 100.0, 10000.0, 105.0, 10500.0, 500.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].holding_type == "invit"

    def test_wrong_header_raises_value_error(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        bad_headers = ["Name", "Code", "Qty", "Price", "Value", "Close", "CVal", "PnL"]
        _create_stocks_xlsx(xlsx, headers=bad_headers)

        with pytest.raises(ValueError, match="Unexpected header"):
            parse_stocks_xlsx(str(xlsx))

    def test_malformed_row_skipped(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Good Stock", "INE002A01018", 10, 100.0, 1000.0, 110.0, 1100.0, 100.0],
            ["Bad Stock", "INE003A01018", "not_a_number", 100.0, 1000.0, 110.0, 1100.0, 100.0],
            ["Another Good", "INE004A01018", 5, 200.0, 1000.0, 210.0, 1050.0, 50.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 2
        assert holdings[0].name == "Good Stock"
        assert holdings[1].name == "Another Good"

    def test_empty_file_returns_empty_list(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        _create_stocks_xlsx(xlsx, rows=[])

        holdings = parse_stocks_xlsx(str(xlsx))
        assert holdings == []

    def test_pnl_percent_computed(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Test Stock", "INE001A01000", 10, 100.0, 1000.0, 120.0, 1200.0, 200.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert holdings[0].pnl_percent == 20.0

    def test_none_row_values_skipped(self, tmp_path):
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            [None, None, None, None, None, None, None, None],
            ["Valid", "INE001A01000", 1, 10.0, 10.0, 11.0, 11.0, 1.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].name == "Valid"

    def test_malformed_row_logs_warning(self, tmp_path, caplog):
        """Malformed rows should emit a warning log with the row number (Req 1.6)."""
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Bad Stock", "INE003A01018", "not_a_number", 100.0, 1000.0, 110.0, 1100.0, 100.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        with caplog.at_level(logging.WARNING):
            holdings = parse_stocks_xlsx(str(xlsx))

        assert len(holdings) == 0
        assert any("Row 12" in msg for msg in caplog.messages)

    def test_mixed_holding_types(self, tmp_path):
        """A file with stock, ETF, and InvIT rows should classify each correctly."""
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Reliance Industries", "INE002A01018", 10, 2500.0, 25000.0, 2600.0, 26000.0, 1000.0],
            ["Motilal Oswal Nasdaq 100", "INF204KB17I5", 20, 100.0, 2000.0, 110.0, 2200.0, 200.0],
            ["Powergrid InvIT", "INE183W23014", 100, 100.0, 10000.0, 105.0, 10500.0, 500.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 3
        assert holdings[0].holding_type == "stock"
        assert holdings[1].holding_type == "etf"
        assert holdings[2].holding_type == "invit"

    def test_default_invit_isins_used(self, tmp_path):
        """parse_stocks_xlsx should use KNOWN_INVIT_ISINS when invit_isins is None."""
        xlsx = tmp_path / "stocks.xlsx"
        for invit_isin in KNOWN_INVIT_ISINS:
            rows = [
                ["InvIT Holding", invit_isin, 50, 100.0, 5000.0, 105.0, 5250.0, 250.0],
            ]
            _create_stocks_xlsx(xlsx, rows=rows)

            holdings = parse_stocks_xlsx(str(xlsx))
            assert len(holdings) == 1
            assert holdings[0].holding_type == "invit"

    def test_etf_name_keyword_in_parsing(self, tmp_path):
        """ETF classification via name keyword within full parse (Req 1.2, 1.5)."""
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Nippon India ETF Nifty 50", "INE123X01000", 15, 200.0, 3000.0, 210.0, 3150.0, 150.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].holding_type == "etf"
        assert holdings[0].name == "Nippon India ETF Nifty 50"

    def test_pnl_percent_zero_buy_value(self, tmp_path):
        """When buy_value is 0, pnl_percent should be 0.0 (no division error)."""
        xlsx = tmp_path / "stocks.xlsx"
        rows = [
            ["Zero Buy", "INE001A01000", 1, 0.0, 0.0, 10.0, 10.0, 10.0],
        ]
        _create_stocks_xlsx(xlsx, rows=rows)

        holdings = parse_stocks_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].pnl_percent == 0.0
