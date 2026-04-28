"""Property-based tests for the Mutual Funds Parser.

**Validates: Requirements 2.2**

Tests Property 6 (MF column extraction completeness) from the design document.
"""

from __future__ import annotations

import os
import tempfile

import openpyxl
from hypothesis import given, settings, assume
from hypothesis import strategies as st

from parsers.groww_mf_parser import parse_mf_xlsx
from parsers.models import MFHolding

# ── XLSX-safe strategies ─────────────────────────────────────────────

# Floats that survive XLSX round-trip without precision loss (2 decimal places).
_xlsx_safe_positive_float = st.floats(
    min_value=0.01, max_value=999_999.99, allow_nan=False, allow_infinity=False,
).map(lambda x: round(x, 2))

_xlsx_safe_any_float = st.floats(
    min_value=-999_999.99, max_value=999_999.99, allow_nan=False, allow_infinity=False,
).map(lambda x: round(x, 2))

# Printable ASCII text safe for openpyxl (no control chars).
_xlsx_safe_text = st.text(
    alphabet=st.sampled_from(
        "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 .-_()"
    ),
    min_size=1,
    max_size=40,
).filter(lambda s: s.strip())

_xlsx_safe_folio = st.text(
    alphabet=st.sampled_from("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789/"),
    min_size=1,
    max_size=20,
).filter(lambda s: s.strip())

_source_strategy = st.sampled_from(["SIP", "Lumpsum", "STP"])

# XIRR as a percentage float >= 1 or <= -1 so the parser keeps it as-is
# (avoids the decimal heuristic where -1 < x < 1 gets multiplied by 100).
_xirr_strategy = st.floats(
    min_value=1.0, max_value=999.99, allow_nan=False, allow_infinity=False,
).map(lambda x: round(x, 2))


# ── Helpers ──────────────────────────────────────────────────────────

_EXPECTED_HEADERS = [
    "Scheme Name", "AMC", "Category", "Sub-category", "Folio No.",
    "Source", "Units", "Invested Value", "Current Value", "Returns", "XIRR",
]


def _create_mf_xlsx(path, rows):
    """Create a Groww-style MF XLSX with correct headers and given data rows.

    Headers go into row 21 (columns A-K). Data rows start at row 23.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    for col_idx, hdr in enumerate(_EXPECTED_HEADERS, start=1):
        ws.cell(row=21, column=col_idx, value=hdr)

    for row_offset, row_data in enumerate(rows):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=23 + row_offset, column=col_idx, value=val)

    wb.save(path)


# ── Property 6: MF column extraction completeness ───────────────────
# **Validates: Requirements 2.2**


@given(
    scheme_name=_xlsx_safe_text,
    amc=_xlsx_safe_text,
    category=_xlsx_safe_text,
    sub_category=_xlsx_safe_text,
    folio_no=_xlsx_safe_folio,
    source=_source_strategy,
    units=_xlsx_safe_positive_float,
    invested_value=_xlsx_safe_positive_float,
    current_value=_xlsx_safe_positive_float,
    returns_absolute=_xlsx_safe_any_float,
    xirr=_xirr_strategy,
)
@settings(max_examples=20, deadline=None)
def test_property_6_mf_column_extraction_completeness(
    scheme_name, amc, category, sub_category, folio_no,
    source, units, invested_value, current_value, returns_absolute, xirr,
):
    """For any valid row of mutual fund data (with all required columns
    present at the expected positions), parse_mf_xlsx should produce an
    MFHolding object containing all 11 required fields: scheme_name, amc,
    category, sub_category, folio_no, source, units, invested_value,
    current_value, returns_absolute, xirr.

    **Validates: Requirements 2.2**
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "mf.xlsx")
        row = [
            scheme_name, amc, category, sub_category, folio_no,
            source, units, invested_value, current_value, returns_absolute, xirr,
        ]
        _create_mf_xlsx(xlsx_path, [row])

        holdings = parse_mf_xlsx(xlsx_path)

        assert len(holdings) == 1
        h = holdings[0]
        assert isinstance(h, MFHolding)

        # All 11 required fields must be present and match input.
        assert h.scheme_name == scheme_name.strip()
        assert h.amc == amc.strip()
        assert h.category == category.strip()
        assert h.sub_category == sub_category.strip()
        assert h.folio_no == folio_no.strip()
        assert h.source == source.strip()
        assert h.units == units
        assert h.invested_value == invested_value
        assert h.current_value == current_value
        assert h.returns_absolute == returns_absolute
        assert h.xirr == xirr
