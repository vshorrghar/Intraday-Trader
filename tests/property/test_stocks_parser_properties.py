"""Property-based tests for the Stocks Parser.

**Validates: Requirements 1.2, 1.3, 1.4, 22.1, 22.2, 22.3, 22.4, 22.5**

Tests Property 4 (Holding classification correctness and exclusivity) and
Property 5 (Stock column extraction completeness) from the design document.
"""

from __future__ import annotations

import openpyxl
from hypothesis import given, settings, assume
from hypothesis import strategies as st

from parsers.groww_stocks_parser import (
    KNOWN_INVIT_ISINS,
    _ETF_NAME_KEYWORDS,
    classify_holding,
    parse_stocks_xlsx,
)
from parsers.models import StockHolding
from tests.conftest import (
    ine_isin_strategy,
    inf_isin_strategy,
    positive_int,
)

# Floats that survive XLSX round-trip without precision loss (2 decimal places).
_xlsx_safe_positive_float = st.floats(
    min_value=0.01, max_value=999_999.99, allow_nan=False, allow_infinity=False,
).map(lambda x: round(x, 2))

_xlsx_safe_any_float = st.floats(
    min_value=-999_999.99, max_value=999_999.99, allow_nan=False, allow_infinity=False,
).map(lambda x: round(x, 2))

# Printable ASCII names safe for openpyxl (no control chars).
_xlsx_safe_name = st.text(
    alphabet=st.sampled_from(
        "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 .-_()"
    ),
    min_size=1,
    max_size=40,
).filter(lambda s: s.strip())

# ── Strategies ───────────────────────────────────────────────────────

# Holding names that do NOT contain any ETF keyword.
_etf_keywords_upper = tuple(k.upper() for k in _ETF_NAME_KEYWORDS)

non_etf_name = (
    st.text(
        alphabet=st.sampled_from("ABCDGHIJKLMOPQRUWXYZabcdghijklmopqruwxyz 0123456789"),
        min_size=1,
        max_size=40,
    )
    .filter(lambda s: s.strip())
    .filter(lambda s: not any(kw in s.upper() for kw in _etf_keywords_upper))
)

# Holding names that DO contain at least one ETF keyword.
etf_keyword_strategy = st.sampled_from(list(_ETF_NAME_KEYWORDS))

etf_name = st.builds(
    lambda prefix, kw, suffix: f"{prefix} {kw} {suffix}",
    prefix=st.text(alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ ", min_size=0, max_size=10),
    kw=etf_keyword_strategy,
    suffix=st.text(alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ 0123456789", min_size=0, max_size=10),
).filter(lambda s: s.strip())

# InvIT ISIN strategy — picks from the known set.
invit_isin_strategy = st.sampled_from(sorted(KNOWN_INVIT_ISINS))


# ── Property 4: Holding classification correctness and exclusivity ───
# **Validates: Requirements 1.2, 1.3, 1.4, 22.1, 22.2, 22.3, 22.4, 22.5**


@given(isin=ine_isin_strategy, name=non_etf_name)
@settings(max_examples=200, deadline=None)
def test_property_4a_ine_isin_not_in_invit_classifies_as_stock(isin, name):
    """ISINs starting with 'INE' that are NOT in the InvIT list should
    classify as 'stock'.

    **Validates: Requirements 1.3, 22.1**
    """
    assume(isin not in KNOWN_INVIT_ISINS)
    result = classify_holding(isin, name, KNOWN_INVIT_ISINS)
    assert result == "stock"


@given(isin=inf_isin_strategy, name=non_etf_name)
@settings(max_examples=200, deadline=None)
def test_property_4b_inf_isin_classifies_as_etf(isin, name):
    """ISINs starting with 'INF' should classify as 'etf'.

    **Validates: Requirements 1.4, 22.2**
    """
    result = classify_holding(isin, name, KNOWN_INVIT_ISINS)
    assert result == "etf"


@given(isin=ine_isin_strategy, name=etf_name)
@settings(max_examples=200, deadline=None)
def test_property_4c_etf_name_keyword_classifies_as_etf(isin, name):
    """Names containing ETF keywords should classify as 'etf' regardless
    of ISIN prefix.

    **Validates: Requirements 1.4, 22.3**
    """
    assume(isin not in KNOWN_INVIT_ISINS)
    result = classify_holding(isin, name, KNOWN_INVIT_ISINS)
    assert result == "etf"


@given(isin=invit_isin_strategy, name=non_etf_name)
@settings(max_examples=50, deadline=None)
def test_property_4d_invit_isin_classifies_as_invit(isin, name):
    """ISINs in the known InvIT list should classify as 'invit'.

    **Validates: Requirements 22.4**
    """
    result = classify_holding(isin, name, KNOWN_INVIT_ISINS)
    assert result == "invit"


@given(
    isin=st.one_of(ine_isin_strategy, inf_isin_strategy, invit_isin_strategy),
    name=st.one_of(non_etf_name, etf_name),
)
@settings(max_examples=300, deadline=None)
def test_property_4e_classification_returns_exactly_one_valid_type(isin, name):
    """classify_holding should always return exactly one of 'stock', 'etf',
    or 'invit' for any ISIN/name combination.

    **Validates: Requirements 22.5**
    """
    result = classify_holding(isin, name, KNOWN_INVIT_ISINS)
    assert result in {"stock", "etf", "invit"}


# ── Property 5: Stock column extraction completeness ─────────────────
# **Validates: Requirements 1.2**


def _create_test_xlsx(path, rows):
    """Create a Groww-style stocks XLSX with correct headers and given data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Stock Name", "ISIN", "Quantity", "Average buy price",
        "Buy value", "Closing price", "Closing value", "Unrealised P&L",
    ]
    for col_idx, hdr in enumerate(headers, start=1):
        ws.cell(row=11, column=col_idx, value=hdr)
    for row_offset, row_data in enumerate(rows):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=12 + row_offset, column=col_idx, value=val)
    wb.save(path)


@given(
    name=_xlsx_safe_name,
    isin=ine_isin_strategy,
    quantity=positive_int,
    avg_buy_price=_xlsx_safe_positive_float,
    buy_value=_xlsx_safe_positive_float,
    closing_price=_xlsx_safe_positive_float,
    closing_value=_xlsx_safe_positive_float,
    unrealised_pnl=_xlsx_safe_any_float,
)
@settings(max_examples=20, deadline=None)
def test_property_5_stock_column_extraction_completeness(
    name, isin, quantity, avg_buy_price, buy_value,
    closing_price, closing_value, unrealised_pnl,
):
    """For any valid row of stock data, parse_stocks_xlsx should produce
    a StockHolding with all 8 required fields populated.

    **Validates: Requirements 1.2**
    """
    import tempfile, os

    assume(isin not in KNOWN_INVIT_ISINS)

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "stocks.xlsx")
        row = [name, isin, quantity, avg_buy_price, buy_value, closing_price, closing_value, unrealised_pnl]
        _create_test_xlsx(xlsx_path, [row])

        holdings = parse_stocks_xlsx(str(xlsx_path))

        assert len(holdings) == 1
        h = holdings[0]
        assert isinstance(h, StockHolding)

        # All 8 required fields must be present and match input.
        assert h.name == name.strip()
        assert h.isin == isin.strip()
        assert h.quantity == quantity
        assert h.avg_buy_price == avg_buy_price
        assert h.buy_value == buy_value
        assert h.groww_closing_price == closing_price
        assert h.groww_closing_value == closing_value
        assert h.unrealised_pnl == unrealised_pnl
