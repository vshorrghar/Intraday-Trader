"""Unit tests for parsers.groww_mf_parser.

Validates: Requirements 2.1, 2.3, 2.4
"""

import logging

import openpyxl
import pytest

from parsers.groww_mf_parser import parse_mf_xlsx


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_DEFAULT_HEADERS = [
    "Scheme Name", "AMC", "Category", "Sub-category", "Folio No.",
    "Source", "Units", "Invested Value", "Current Value", "Returns", "XIRR",
]


def _create_mf_xlsx(path, headers=None, rows=None):
    """Create a minimal Groww-style MF XLSX at *path*.

    Headers go into row 21 (columns A-K).  Data rows start at row 23.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    if headers is None:
        headers = _DEFAULT_HEADERS

    for col_idx, hdr in enumerate(headers, start=1):  # A=1
        ws.cell(row=21, column=col_idx, value=hdr)

    if rows:
        for row_offset, row_data in enumerate(rows):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=23 + row_offset, column=col_idx, value=val)

    wb.save(path)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestParseMfXlsx:
    """Tests for parse_mf_xlsx()."""

    def test_valid_parsing(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Axis Bluechip Fund", "Axis", "Equity", "Large Cap",
                "12345", "SIP", 100.5, 50000.0, 55000.0, 5000.0, "11.87%",
            ],
            [
                "HDFC Mid-Cap Opportunities", "HDFC", "Equity", "Mid Cap",
                "67890", "Lumpsum", 200.0, 80000.0, 90000.0, 10000.0, 0.1520,
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert len(holdings) == 2

        h0 = holdings[0]
        assert h0.scheme_name == "Axis Bluechip Fund"
        assert h0.amc == "Axis"
        assert h0.category == "Equity"
        assert h0.sub_category == "Large Cap"
        assert h0.folio_no == "12345"
        assert h0.source == "SIP"
        assert h0.units == 100.5
        assert h0.invested_value == 50000.0
        assert h0.current_value == 55000.0
        assert h0.returns_absolute == 5000.0
        assert h0.xirr == 11.87
        assert h0.returns_percent == 10.0  # 5000/50000*100

    def test_xirr_as_decimal_float(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "SBI Fund", "SBI", "Equity", "Large Cap",
                "11111", "SIP", 50.0, 10000.0, 11000.0, 1000.0, 0.1187,
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert holdings[0].xirr == 11.87

    def test_xirr_as_percentage_string(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "SBI Fund", "SBI", "Equity", "Large Cap",
                "11111", "SIP", 50.0, 10000.0, 11000.0, 1000.0, "25.5%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert holdings[0].xirr == 25.5

    def test_xirr_as_large_float(self, tmp_path):
        """XIRR values >= 1 or <= -1 are kept as-is (already percentages)."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "SBI Fund", "SBI", "Equity", "Large Cap",
                "11111", "SIP", 50.0, 10000.0, 11000.0, 1000.0, 25.5,
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert holdings[0].xirr == 25.5

    def test_wrong_header_raises_value_error(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        bad_headers = [
            "Name", "Company", "Cat", "SubCat", "Folio",
            "Src", "Qty", "Invested", "Current", "Ret", "IRR",
        ]
        _create_mf_xlsx(xlsx, headers=bad_headers)

        with pytest.raises(ValueError, match="Unexpected header"):
            parse_mf_xlsx(str(xlsx))

    def test_malformed_row_skipped(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Good Fund", "AMC1", "Equity", "Large Cap",
                "111", "SIP", 10.0, 1000.0, 1100.0, 100.0, "10%",
            ],
            [
                "Bad Fund", "AMC2", "Equity", "Mid Cap",
                "222", "SIP", "not_a_number", 1000.0, 1100.0, 100.0, "10%",
            ],
            [
                "Another Good", "AMC3", "Debt", "Short Duration",
                "333", "Lumpsum", 20.0, 2000.0, 2200.0, 200.0, "8%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert len(holdings) == 2
        assert holdings[0].scheme_name == "Good Fund"
        assert holdings[1].scheme_name == "Another Good"

    def test_empty_file_returns_empty_list(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        _create_mf_xlsx(xlsx, rows=[])

        holdings = parse_mf_xlsx(str(xlsx))
        assert holdings == []

    def test_none_row_values_skipped(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [None, None, None, None, None, None, None, None, None, None, None],
            [
                "Valid Fund", "AMC", "Equity", "Large Cap",
                "999", "SIP", 5.0, 500.0, 550.0, 50.0, "12%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].scheme_name == "Valid Fund"

    def test_returns_percent_computed(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Test Fund", "AMC", "Equity", "Large Cap",
                "100", "SIP", 10.0, 10000.0, 12000.0, 2000.0, "15%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert holdings[0].returns_percent == 20.0  # 2000/10000*100

    def test_returns_percent_zero_when_no_investment(self, tmp_path):
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Zero Fund", "AMC", "Equity", "Large Cap",
                "100", "SIP", 10.0, 0.0, 0.0, 0.0, "0%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert holdings[0].returns_percent == 0.0

    def test_all_11_columns_extracted(self, tmp_path):
        """Verify all 11 required fields are populated from the XLSX."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Scheme A", "AMC A", "Equity", "Multi Cap",
                "FOL001", "SIP", 150.75, 75000.0, 82000.0, 7000.0, "14.2%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        h = holdings[0]
        assert h.scheme_name == "Scheme A"
        assert h.amc == "AMC A"
        assert h.category == "Equity"
        assert h.sub_category == "Multi Cap"
        assert h.folio_no == "FOL001"
        assert h.source == "SIP"
        assert h.units == 150.75
        assert h.invested_value == 75000.0
        assert h.current_value == 82000.0
        assert h.returns_absolute == 7000.0
        assert h.xirr == 14.2

    def test_malformed_row_logs_warning(self, tmp_path, caplog):
        """Malformed rows emit a warning log with the row number (Req 2.4)."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Bad Fund", "AMC", "Equity", "Large Cap",
                "111", "SIP", None, 1000.0, 1100.0, 100.0, "10%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        with caplog.at_level(logging.WARNING, logger="parsers.groww_mf_parser"):
            holdings = parse_mf_xlsx(str(xlsx))

        assert len(holdings) == 0
        assert any("skipping" in rec.message.lower() for rec in caplog.records)

    def test_empty_scheme_name_skipped_with_log(self, tmp_path, caplog):
        """Rows with empty scheme name are skipped and logged (Req 2.4)."""
        xlsx = tmp_path / "mf.xlsx"
        # Write a row where scheme_name cell is a whitespace-only string
        rows = [
            [
                "  ", "AMC", "Equity", "Large Cap",
                "111", "SIP", 10.0, 1000.0, 1100.0, 100.0, "10%",
            ],
            [
                "Valid Fund", "AMC2", "Debt", "Short Duration",
                "222", "Lumpsum", 20.0, 2000.0, 2200.0, 200.0, "8%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        with caplog.at_level(logging.WARNING, logger="parsers.groww_mf_parser"):
            holdings = parse_mf_xlsx(str(xlsx))

        assert len(holdings) == 1
        assert holdings[0].scheme_name == "Valid Fund"
        assert any("empty scheme name" in rec.message.lower() for rec in caplog.records)

    def test_empty_amc_skipped(self, tmp_path):
        """Rows with empty AMC are skipped (Req 2.4)."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Some Fund", "  ", "Equity", "Large Cap",
                "111", "SIP", 10.0, 1000.0, 1100.0, 100.0, "10%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert len(holdings) == 0

    def test_comma_in_invested_value(self, tmp_path):
        """Invested/current values with commas (e.g. '50,000') are parsed correctly."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Comma Fund", "AMC", "Equity", "Large Cap",
                "111", "SIP", 10.0, "50,000", "55,000", 5000.0, "12%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert len(holdings) == 1
        assert holdings[0].invested_value == 50000.0
        assert holdings[0].current_value == 55000.0

    def test_file_not_found_raises_error(self):
        """Non-existent file path raises FileNotFoundError (Req 2.1)."""
        with pytest.raises(FileNotFoundError):
            parse_mf_xlsx("/nonexistent/path/mf.xlsx")

    def test_wrong_header_error_message_contains_expected_and_actual(self, tmp_path):
        """ValueError message includes both expected and actual headers (Req 2.3)."""
        xlsx = tmp_path / "mf.xlsx"
        bad_headers = [
            "Name", "Company", "Cat", "SubCat", "Folio",
            "Src", "Qty", "Invested", "Current", "Ret", "IRR",
        ]
        _create_mf_xlsx(xlsx, headers=bad_headers)

        with pytest.raises(ValueError, match="Expected") as exc_info:
            parse_mf_xlsx(str(xlsx))
        # Verify the error message contains both expected and actual header info
        msg = str(exc_info.value)
        assert "Scheme Name" in msg  # part of expected headers
        assert "Name" in msg  # part of actual headers

    def test_malformed_row_log_contains_row_number(self, tmp_path, caplog):
        """Log message for malformed rows includes the row number (Req 2.4)."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Good Fund", "AMC", "Equity", "Large Cap",
                "111", "SIP", 10.0, 1000.0, 1100.0, 100.0, "10%",
            ],
            [
                "Bad Fund", "AMC", "Equity", "Mid Cap",
                "222", "SIP", None, 1000.0, 1100.0, 100.0, "10%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        with caplog.at_level(logging.WARNING, logger="parsers.groww_mf_parser"):
            holdings = parse_mf_xlsx(str(xlsx))

        assert len(holdings) == 1
        # Row 24 is the second data row (row 23 = first data row)
        assert any("24" in rec.message for rec in caplog.records)

    def test_multiple_valid_rows_preserve_order(self, tmp_path):
        """Multiple valid rows are returned in the same order as the XLSX (Req 2.1)."""
        xlsx = tmp_path / "mf.xlsx"
        rows = [
            [
                "Fund Alpha", "AMC1", "Equity", "Large Cap",
                "001", "SIP", 10.0, 1000.0, 1100.0, 100.0, "10%",
            ],
            [
                "Fund Beta", "AMC2", "Debt", "Short Duration",
                "002", "Lumpsum", 20.0, 2000.0, 2200.0, 200.0, "8%",
            ],
            [
                "Fund Gamma", "AMC3", "Hybrid", "Balanced",
                "003", "STP", 30.0, 3000.0, 3300.0, 300.0, "15%",
            ],
        ]
        _create_mf_xlsx(xlsx, rows=rows)

        holdings = parse_mf_xlsx(str(xlsx))
        assert len(holdings) == 3
        assert holdings[0].scheme_name == "Fund Alpha"
        assert holdings[1].scheme_name == "Fund Beta"
        assert holdings[2].scheme_name == "Fund Gamma"
