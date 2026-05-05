"""Unit tests for parsers.groww_pnl_parser."""

from datetime import datetime

import openpyxl
import pytest

from parsers.groww_pnl_parser import (
    classify_tax_term,
    compute_holding_period,
    parse_pnl_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_pnl_xlsx(path, trade_rows=None, scrip_rows=None, sheets=None):
    """Create a minimal Groww-style P&L XLSX at *path*.

    By default creates both "Trade Level" and "Scrip Level" sheets.
    Pass *sheets* to override which sheets are created.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet.
    wb.remove(wb.active)

    if sheets is None:
        sheets = ["Trade Level", "Scrip Level"]

    if "Trade Level" in sheets:
        ws_trade = wb.create_sheet("Trade Level")
        # Header row 1.
        trade_headers = [
            "Stock Name", "ISIN", "Quantity", "Buy Date",
            "Buy Price", "Closing Date", "Closing Price", "Unrealised P&L",
        ]
        for col_idx, hdr in enumerate(trade_headers, start=1):
            ws_trade.cell(row=1, column=col_idx, value=hdr)
        if trade_rows:
            for row_offset, row_data in enumerate(trade_rows):
                for col_idx, val in enumerate(row_data, start=1):
                    ws_trade.cell(row=2 + row_offset, column=col_idx, value=val)

    if "Scrip Level" in sheets:
        ws_scrip = wb.create_sheet("Scrip Level")
        scrip_headers = [
            "Stock Name", "ISIN", "Buy Quantity", "Buy Avg Price",
            "Sell Quantity", "Sell Avg Price", "Realised P&L",
        ]
        for col_idx, hdr in enumerate(scrip_headers, start=1):
            ws_scrip.cell(row=1, column=col_idx, value=hdr)
        if scrip_rows:
            for row_offset, row_data in enumerate(scrip_rows):
                for col_idx, val in enumerate(row_data, start=1):
                    ws_scrip.cell(row=2 + row_offset, column=col_idx, value=val)

    wb.save(path)


# ---------------------------------------------------------------------------
# compute_holding_period tests
# ---------------------------------------------------------------------------


class TestComputeHoldingPeriod:
    """Tests for compute_holding_period()."""

    def test_same_day(self):
        d = datetime(2024, 1, 15)
        assert compute_holding_period(d, d) == 0

    def test_one_day(self):
        buy = datetime(2024, 1, 15)
        cur = datetime(2024, 1, 16)
        assert compute_holding_period(buy, cur) == 1

    def test_one_year(self):
        buy = datetime(2023, 1, 1)
        cur = datetime(2024, 1, 1)
        assert compute_holding_period(buy, cur) == 365

    def test_exact_365_days(self):
        buy = datetime(2023, 6, 1)
        cur = datetime(2024, 5, 31)
        assert compute_holding_period(buy, cur) == 365

    def test_large_period(self):
        buy = datetime(2020, 1, 1)
        cur = datetime(2024, 1, 1)
        assert compute_holding_period(buy, cur) == 1461  # 4 years incl leap


# ---------------------------------------------------------------------------
# classify_tax_term tests
# ---------------------------------------------------------------------------


class TestClassifyTaxTerm:
    """Tests for classify_tax_term()."""

    def test_stock_short_term_below_365(self):
        assert classify_tax_term(364, "stock") == "short_term"

    def test_stock_long_term_at_365(self):
        assert classify_tax_term(365, "stock") == "long_term"

    def test_stock_long_term_above_365(self):
        assert classify_tax_term(400, "stock") == "long_term"

    def test_equity_mf_short_term_below_365(self):
        assert classify_tax_term(364, "equity_mf") == "short_term"

    def test_equity_mf_long_term_at_365(self):
        assert classify_tax_term(365, "equity_mf") == "long_term"

    def test_debt_mf_short_term_below_1095(self):
        assert classify_tax_term(1094, "debt_mf") == "short_term"

    def test_debt_mf_long_term_at_1095(self):
        assert classify_tax_term(1095, "debt_mf") == "long_term"

    def test_debt_mf_long_term_above_1095(self):
        assert classify_tax_term(1200, "debt_mf") == "long_term"

    def test_zero_days_is_short_term(self):
        assert classify_tax_term(0, "stock") == "short_term"
        assert classify_tax_term(0, "debt_mf") == "short_term"

    def test_stock_at_365_boundary(self):
        """Exactly 365 days should be long_term for stocks."""
        assert classify_tax_term(365, "stock") == "long_term"

    def test_debt_mf_at_1095_boundary(self):
        """Exactly 1095 days should be long_term for debt MF."""
        assert classify_tax_term(1095, "debt_mf") == "long_term"


# ---------------------------------------------------------------------------
# parse_pnl_xlsx tests
# ---------------------------------------------------------------------------


class TestParsePnlXlsx:
    """Tests for parse_pnl_xlsx()."""

    def test_valid_parsing_both_sheets(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        buy_date = datetime(2023, 6, 15)
        current = datetime(2024, 6, 15)

        trade_rows = [
            ["Reliance", "INE002A01018", 10, buy_date, 2500.0,
             current, 2600.0, 1000.0],
            ["HDFC Bank", "INE040A01034", 5, buy_date, 1600.0,
             current, 1650.0, 250.0],
        ]
        scrip_rows = [
            ["Reliance", "INE002A01018", 10, 2500.0, 0, 0.0, 0.0],
            ["HDFC Bank", "INE040A01034", 5, 1600.0, 2, 1700.0, 200.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows, scrip_rows=scrip_rows)

        trades, scrips = parse_pnl_xlsx(str(xlsx), current_date=current)

        assert len(trades) == 2
        assert trades[0].symbol == "Reliance"
        assert trades[0].isin == "INE002A01018"
        assert trades[0].quantity == 10
        assert trades[0].trade_date == buy_date
        assert trades[0].price == 2500.0

        assert len(scrips) == 2
        assert scrips[0].symbol == "Reliance"
        assert scrips[0].isin == "INE002A01018"
        assert scrips[0].buy_quantity == 10
        assert scrips[0].buy_avg_price == 2500.0
        # 366 days (2024 is leap year)
        assert scrips[0].holding_period_days == 366
        assert scrips[0].tax_classification == "long_term"

    def test_missing_trade_level_sheet(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        _create_pnl_xlsx(xlsx, sheets=["Scrip Level"])

        with pytest.raises(ValueError, match="Trade Level"):
            parse_pnl_xlsx(str(xlsx))

    def test_missing_scrip_level_sheet(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        _create_pnl_xlsx(xlsx, sheets=["Trade Level"])

        with pytest.raises(ValueError, match="Scrip Level"):
            parse_pnl_xlsx(str(xlsx))

    def test_missing_both_sheets(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx)

        with pytest.raises(ValueError, match="Missing required sheet"):
            parse_pnl_xlsx(str(xlsx))

    def test_empty_sheets_return_empty_lists(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        _create_pnl_xlsx(xlsx, trade_rows=[], scrip_rows=[])

        trades, scrips = parse_pnl_xlsx(str(xlsx))
        assert trades == []
        assert scrips == []

    def test_malformed_trade_row_skipped(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        buy_date = datetime(2024, 1, 1)
        current = datetime(2024, 6, 1)

        trade_rows = [
            ["Good Stock", "INE001A01000", 10, buy_date, 100.0,
             current, 110.0, 100.0],
            ["Bad Stock", "INE002A01000", "not_int", buy_date, 100.0,
             current, 110.0, 100.0],
            ["Another Good", "INE003A01000", 5, buy_date, 200.0,
             current, 210.0, 50.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows)

        trades, _ = parse_pnl_xlsx(str(xlsx), current_date=current)
        assert len(trades) == 2
        assert trades[0].symbol == "Good Stock"
        assert trades[1].symbol == "Another Good"

    def test_malformed_scrip_row_skipped(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        current = datetime(2024, 6, 1)

        scrip_rows = [
            ["Good Scrip", "INE001A01000", 10, 100.0, 0, 0.0, 0.0],
            ["Bad Scrip", "INE002A01000", "bad", 100.0, 0, 0.0, 0.0],
            ["Another Good", "INE003A01000", 5, 200.0, 2, 210.0, 20.0],
        ]
        _create_pnl_xlsx(xlsx, scrip_rows=scrip_rows)

        _, scrips = parse_pnl_xlsx(str(xlsx), current_date=current)
        assert len(scrips) == 2
        assert scrips[0].symbol == "Good Scrip"
        assert scrips[1].symbol == "Another Good"

    def test_scrip_uses_trade_buy_date_for_holding_period(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        buy_date = datetime(2023, 1, 1)
        current = datetime(2024, 1, 1)

        trade_rows = [
            ["Stock A", "INE001A01000", 10, buy_date, 100.0,
             current, 110.0, 100.0],
        ]
        scrip_rows = [
            ["Stock A", "INE001A01000", 10, 100.0, 0, 0.0, 0.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows, scrip_rows=scrip_rows)

        _, scrips = parse_pnl_xlsx(str(xlsx), current_date=current)
        assert scrips[0].holding_period_days == 365
        assert scrips[0].tax_classification == "long_term"

    def test_debt_mf_tax_classification(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        buy_date = datetime(2022, 1, 1)
        current = datetime(2024, 12, 31)

        trade_rows = [
            ["Debt Fund", "INE001A01000", 100, buy_date, 10.0,
             current, 11.0, 100.0],
        ]
        scrip_rows = [
            ["Debt Fund", "INE001A01000", 100, 10.0, 0, 0.0, 0.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows, scrip_rows=scrip_rows)

        _, scrips = parse_pnl_xlsx(
            str(xlsx), current_date=current, security_type="debt_mf"
        )
        # 2022-01-01 to 2024-12-31 = 1095 days (exactly at threshold)
        assert scrips[0].holding_period_days == 1095
        assert scrips[0].tax_classification == "long_term"

    def test_debt_mf_short_term(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        buy_date = datetime(2023, 1, 1)
        current = datetime(2024, 12, 30)

        trade_rows = [
            ["Debt Fund", "INE001A01000", 100, buy_date, 10.0,
             current, 11.0, 100.0],
        ]
        scrip_rows = [
            ["Debt Fund", "INE001A01000", 100, 10.0, 0, 0.0, 0.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows, scrip_rows=scrip_rows)

        _, scrips = parse_pnl_xlsx(
            str(xlsx), current_date=current, security_type="debt_mf"
        )
        # 2023-01-01 to 2024-12-30 = 729 days (below 1095)
        assert scrips[0].tax_classification == "short_term"

    def test_none_rows_skipped(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        current = datetime(2024, 6, 1)

        trade_rows = [
            [None, None, None, None, None, None, None, None],
            ["Valid", "INE001A01000", 1, datetime(2024, 1, 1), 10.0,
             current, 11.0, 1.0],
        ]
        scrip_rows = [
            [None, None, None, None, None, None, None],
            ["Valid", "INE001A01000", 1, 10.0, 0, 0.0, 0.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows, scrip_rows=scrip_rows)

        trades, scrips = parse_pnl_xlsx(str(xlsx), current_date=current)
        assert len(trades) == 1
        assert len(scrips) == 1

    def test_trade_date_parsed_as_datetime(self, tmp_path):
        xlsx = tmp_path / "pnl.xlsx"
        buy_date = datetime(2024, 3, 15)
        current = datetime(2024, 6, 15)

        trade_rows = [
            ["Stock", "INE001A01000", 10, buy_date, 100.0,
             current, 110.0, 100.0],
        ]
        _create_pnl_xlsx(xlsx, trade_rows=trade_rows)

        trades, _ = parse_pnl_xlsx(str(xlsx), current_date=current)
        assert isinstance(trades[0].trade_date, datetime)
        assert trades[0].trade_date == buy_date
