"""Property-based tests for the P&L Parser.

**Validates: Requirements 3.2, 3.3, 3.4**

Tests Property 7 (Trade record extraction), Property 8 (Holding period
computation), and Property 9 (Tax term classification) from the design
document.
"""

from __future__ import annotations

import os
import tempfile
from datetime import datetime

import openpyxl
from hypothesis import given, settings, assume
from hypothesis import strategies as st

from parsers.groww_pnl_parser import (
    classify_tax_term,
    compute_holding_period,
    parse_pnl_xlsx,
)
from parsers.models import TradeRecord
from tests.conftest import (
    isin_strategy,
    positive_float,
    positive_int,
    reasonable_datetime,
)

# ── Helpers ──────────────────────────────────────────────────────────

# Printable ASCII symbols safe for openpyxl (no control chars).
_xlsx_safe_symbol = st.text(
    alphabet=st.sampled_from(
        "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 .-_()"
    ),
    min_size=1,
    max_size=30,
).filter(lambda s: s.strip())

# Floats that survive XLSX round-trip without precision loss (2 decimal places).
_xlsx_safe_positive_float = st.floats(
    min_value=0.01, max_value=999_999.99, allow_nan=False, allow_infinity=False,
).map(lambda x: round(x, 2))


def _create_pnl_xlsx(path, trade_rows, scrip_rows=None):
    """Create a Groww-style P&L XLSX with Trade Level and Scrip Level sheets."""
    wb = openpyxl.Workbook()

    # Trade Level sheet
    ws_trade = wb.active
    ws_trade.title = "Trade Level"
    trade_headers = [
        "Stock Name", "ISIN", "Quantity", "Buy Date",
        "Buy Price", "Closing Date", "Closing Price", "Unrealised P&L",
    ]
    for col_idx, hdr in enumerate(trade_headers, start=1):
        ws_trade.cell(row=1, column=col_idx, value=hdr)
    for row_offset, row_data in enumerate(trade_rows):
        for col_idx, val in enumerate(row_data, start=1):
            ws_trade.cell(row=2 + row_offset, column=col_idx, value=val)

    # Scrip Level sheet
    ws_scrip = wb.create_sheet("Scrip Level")
    scrip_headers = [
        "Stock Name", "ISIN", "Buy Quantity", "Buy Avg Price",
        "Sell Quantity", "Sell Avg Price", "Realised P&L",
    ]
    for col_idx, hdr in enumerate(scrip_headers, start=1):
        ws_scrip.cell(row=1, column=col_idx, value=hdr)
    if scrip_rows:
        for row_offset, row_data in enumerate(scrip_rows):
            for col_idx, val in enumerate(row_data, start=1):
                ws_scrip.cell(row=2 + row_offset, column=col_idx, value=val)

    wb.save(path)


# ── Property 7: Trade record extraction ──────────────────────────────
# **Validates: Requirements 3.2**


@given(
    symbol=_xlsx_safe_symbol,
    isin=isin_strategy,
    quantity=positive_int,
    buy_date=reasonable_datetime,
    buy_price=_xlsx_safe_positive_float,
)
@settings(max_examples=30, deadline=None)
def test_property_7_trade_record_extraction(
    symbol, isin, quantity, buy_date, buy_price,
):
    """For any valid trade-level row in a P&L XLSX, the parser should
    extract a TradeRecord with a valid buy date (parseable datetime) and
    a non-empty ISIN string.

    **Validates: Requirements 3.2**
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "pnl.xlsx")
        trade_row = [symbol, isin, quantity, buy_date, buy_price, None, None, None]
        _create_pnl_xlsx(xlsx_path, [trade_row])

        trade_records, _ = parse_pnl_xlsx(xlsx_path)

        assert len(trade_records) == 1
        tr = trade_records[0]
        assert isinstance(tr, TradeRecord)

        # ISIN must be non-empty
        assert isinstance(tr.isin, str)
        assert len(tr.isin) > 0

        # trade_date must be a valid datetime
        assert isinstance(tr.trade_date, datetime)


# ── Property 8: Holding period computation ───────────────────────────
# **Validates: Requirements 3.3**


# Strategy: generate two datetimes where buy_date <= current_date.
_ordered_date_pair = st.tuples(reasonable_datetime, reasonable_datetime).map(
    lambda pair: (min(pair), max(pair))
)


@given(date_pair=_ordered_date_pair)
@settings(max_examples=200, deadline=None)
def test_property_8_holding_period_computation(date_pair):
    """For any two dates where buy_date <= current_date,
    compute_holding_period should return a non-negative integer equal to
    the number of days between the two dates.

    **Validates: Requirements 3.3**
    """
    buy_date, current_date = date_pair

    result = compute_holding_period(buy_date, current_date)

    # Must be a non-negative integer
    assert isinstance(result, int)
    assert result >= 0

    # Must equal the actual day difference
    expected = (current_date - buy_date).days
    assert result == expected


# ── Property 9: Tax term classification ──────────────────────────────
# **Validates: Requirements 3.4**


_security_type_strategy = st.sampled_from(["stock", "equity_mf", "debt_mf"])
_holding_period_strategy = st.integers(min_value=0, max_value=5000)


@given(
    holding_period_days=_holding_period_strategy,
    security_type=_security_type_strategy,
)
@settings(max_examples=300, deadline=None)
def test_property_9_tax_term_classification(holding_period_days, security_type):
    """For any holding period in days and security type, classify_tax_term
    should return 'short_term' when below threshold (365 days for
    stocks/equity MF, 1095 days for debt MF) and 'long_term' when at or
    above.

    **Validates: Requirements 3.4**
    """
    result = classify_tax_term(holding_period_days, security_type)

    # Determine expected threshold
    if security_type == "debt_mf":
        threshold = 1095
    else:
        threshold = 365

    if holding_period_days >= threshold:
        assert result == "long_term", (
            f"Expected 'long_term' for {holding_period_days} days "
            f"({security_type}, threshold={threshold}), got '{result}'"
        )
    else:
        assert result == "short_term", (
            f"Expected 'short_term' for {holding_period_days} days "
            f"({security_type}, threshold={threshold}), got '{result}'"
        )

    # Result must always be one of the two valid values
    assert result in {"short_term", "long_term"}
