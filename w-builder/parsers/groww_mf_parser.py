"""Parser for Groww Mutual Funds Holdings XLSX exports.

Reads the Groww broker mutual funds holdings XLSX file with header at row 23
and data rows starting from row 24. Columns B through L contain:
Scheme Name, AMC, Category, Sub-category, Folio No, Source, Units,
Invested Value, Current Value, Returns, XIRR.

XIRR values may arrive as a percentage string (e.g. "11.87%") or as a
decimal float (e.g. 0.1187). Both forms are normalised to a float
percentage (11.87).
"""

from __future__ import annotations

import logging

import openpyxl

from parsers.models import MFHolding

logger = logging.getLogger(__name__)

# Expected header labels — Groww may use "Folio No." with a period
_EXPECTED_HEADERS: tuple[str, ...] = (
    "Scheme Name",
    "AMC",
    "Category",
    "Sub-category",
    "Folio No.",
    "Source",
    "Units",
    "Invested Value",
    "Current Value",
    "Returns",
    "XIRR",
)

_HEADER_ROW = 21   # Actual Groww file has headers at row 21
_DATA_START_ROW = 23  # Data starts at row 23 (row 22 is blank)
_COL_START = 1   # Column A (1-indexed) — actual Groww file starts at A
_COL_END = 11    # Column K (1-indexed, inclusive)


def _parse_xirr(value) -> float:
    """Normalise an XIRR cell value to a float percentage.

    Handles:
    - Percentage strings like ``"11.87%"`` → ``11.87``
    - Decimal floats like ``0.1187`` → ``11.87``
    - Already-percentage floats like ``11.87`` → ``11.87``
    """
    if isinstance(value, str):
        stripped = value.strip().rstrip("%")
        return float(stripped)

    num = float(value)
    # Heuristic: values in (-1, 1) exclusive are treated as decimal fractions.
    if -1 < num < 1 and num != 0:
        return round(num * 100, 2)
    return num


def parse_mf_xlsx(file_path: str) -> list[MFHolding]:
    """Parse a Groww Mutual Funds Holdings XLSX file.

    Parameters
    ----------
    file_path:
        Path to the XLSX file exported from Groww.

    Returns
    -------
    list[MFHolding]
        Parsed mutual fund holdings.

    Raises
    ------
    ValueError
        If the header row (row 23) does not match the expected structure.
    FileNotFoundError
        If *file_path* does not exist.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # --- Validate header at row 23 (columns B-L) -----------------------
    header_cells = [
        ws.cell(row=_HEADER_ROW, column=col).value
        for col in range(_COL_START, _COL_END + 1)
    ]
    header_values = tuple(
        str(v).strip() if v is not None else "" for v in header_cells
    )

    if header_values != _EXPECTED_HEADERS:
        wb.close()
        raise ValueError(
            f"Unexpected header at row {_HEADER_ROW}. "
            f"Expected {_EXPECTED_HEADERS}, got {header_values}"
        )

    # --- Parse data rows from row 24 onward -----------------------------
    holdings: list[MFHolding] = []

    current_row = _DATA_START_ROW
    for row in ws.iter_rows(
        min_row=_DATA_START_ROW,
        min_col=_COL_START,
        max_col=_COL_END,
        values_only=False,
    ):
        row_num = current_row
        current_row += 1

        raw = [cell.value for cell in row]

        # Skip completely empty rows.
        if all(v is None for v in raw):
            continue

        try:
            scheme_name = str(raw[0]).strip()
            amc = str(raw[1]).strip()
            category = str(raw[2]).strip()
            sub_category = str(raw[3]).strip()
            folio_no = str(raw[4]).strip()
            source = str(raw[5]).strip()
            units = float(raw[6])
            invested_value = float(str(raw[7]).replace(',', ''))
            current_value = float(str(raw[8]).replace(',', ''))
            returns_absolute = float(raw[9])
            xirr = _parse_xirr(raw[10])
        except (TypeError, ValueError) as exc:
            logger.warning(
                "Row %d: skipping due to missing/malformed value – %s",
                row_num,
                exc,
            )
            continue

        if not scheme_name or not amc:
            logger.warning(
                "Row %d: skipping due to empty scheme name or AMC", row_num
            )
            continue

        returns_percent = (
            (returns_absolute / invested_value * 100)
            if invested_value > 0
            else 0.0
        )

        holdings.append(
            MFHolding(
                scheme_name=scheme_name,
                amc=amc,
                category=category,
                sub_category=sub_category,
                folio_no=folio_no,
                source=source,
                units=units,
                invested_value=invested_value,
                current_value=current_value,
                returns_absolute=returns_absolute,
                xirr=xirr,
                returns_percent=round(returns_percent, 2),
            )
        )

    wb.close()
    return holdings
