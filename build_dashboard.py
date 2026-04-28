#!/usr/bin/env python3
"""Build dashboard data from latest analysis + local parsed portfolio.

Merges Claude analysis JSON with locally parsed stock/MF data so the
dashboard always shows ALL holdings even if Claude only analyzed a subset.
"""
import json, os, sys, math
from pathlib import Path
from datetime import datetime

print("🔧 Building dashboard data...")

# Get XLSX file dates
import re
from pathlib import Path

def get_xlsx_date(filepath):
    """Extract date from Groww XLSX filename or fall back to file modification time."""
    name = os.path.basename(filepath)
    dates = re.findall(r'(\d{2}-\d{2}-\d{4})', name)
    if dates:
        try:
            return max(datetime.strptime(d, '%d-%m-%Y') for d in dates).strftime('%d-%b-%Y')
        except ValueError:
            pass
    # Fallback: file modification time
    if os.path.exists(filepath):
        mtime = os.path.getmtime(filepath)
        return datetime.fromtimestamp(mtime).strftime('%d-%b-%Y')
    return None

xlsx_dates = {}
for label, path in [
    ('stocks', 'input/Stocks_Holdings_Statement.xlsx'),
    ('mf', 'input/Mutual_Funds.xlsx'),
    ('pnl', 'input/Stocks_PnL_Report.xlsx'),
]:
    if os.path.exists(path):
        xlsx_dates[label] = get_xlsx_date(path)

# Load Claude analysis if available
analysis = {}
for f in ['output/reports/latest.json', 'output/latest_analysis.json']:
    if os.path.exists(f):
        with open(f) as fh:
            analysis = json.load(fh)
        print(f"  ✅ Loaded analysis from {f}")
        break

# Deep find helper
def fk(o, k):
    if not o or not isinstance(o, dict): return None
    if k in o: return o[k]
    for v in o.values():
        r = fk(v, k)
        if r is not None: return r
    return None

# Parse local portfolio for complete data
stocks, mfs = [], []
try:
    from parsers.groww_stocks_parser import parse_stocks_xlsx
    from parsers.groww_mf_parser import parse_mf_xlsx
    if os.path.exists('input/Stocks_Holdings_Statement.xlsx'):
        stocks = parse_stocks_xlsx('input/Stocks_Holdings_Statement.xlsx')
        print(f"  ✅ Parsed {len(stocks)} stocks from XLSX")
    if os.path.exists('input/Mutual_Funds.xlsx'):
        mfs = parse_mf_xlsx('input/Mutual_Funds.xlsx')
        print(f"  ✅ Parsed {len(mfs)} MF schemes from XLSX")
except Exception as e:
    print(f"  ⚠️ Parser error: {e}")

# Build verdicts for ALL stocks (merge Claude verdicts with parsed data)
claude_verdicts = fk(analysis, 'individual_stock_verdicts') or fk(analysis, 'stock_verdicts') or []
verdict_map = {v.get('name',''): v for v in claude_verdicts}

all_stock_verdicts = []
stock_by_name = {h.name: h for h in stocks}
for h in stocks:
    cv = verdict_map.get(h.name)
    base = {
        'qty': h.quantity,
        'avg_price': round(h.avg_buy_price, 1),
        'current_price': round(h.groww_closing_price, 1),
        'current_value': round(h.groww_closing_value, 0),
        'pnl_pct': h.pnl_percent,
        'pnl_abs': round(h.unrealised_pnl, 0),
        'holding_type': h.holding_type,
    }
    if cv:
        cv.update(base)
        all_stock_verdicts.append(cv)
    else:
        # No Claude verdict — generate smart one from P&L
        verdict = 'HOLD'
        action = ''
        if h.pnl_percent <= -50:
            verdict, action = 'EXIT', 'Exit — down >50%, no recovery thesis'
        elif h.pnl_percent <= -30:
            verdict, action = 'SELL', 'Consider selling — significant loss'
        elif h.pnl_percent >= 100:
            verdict, action = 'HOLD', 'Book 30-50% profits — up >100%'
        elif h.pnl_percent >= 50:
            verdict, action = 'HOLD', 'Strong performer — trail stop loss'
        elif 10 <= h.pnl_percent < 50 and h.groww_closing_value >= 5000:
            verdict, action = 'BUY', 'Momentum stock — consider adding on dips'
        elif -10 <= h.pnl_percent < 10 and h.groww_closing_value >= 10000:
            verdict, action = 'BUY', 'Consolidating near buy price — good entry zone'
        elif -20 <= h.pnl_percent < -10 and h.groww_closing_value >= 5000:
            verdict, action = 'HOLD', 'Monitor — averaging down risky without thesis'
        else:
            action = 'Monitor — small position'

        all_stock_verdicts.append({
            'name': h.name, 'verdict': verdict, 'action': action,
            'reason': f'P&L: {h.pnl_percent:+.1f}% | Value: ₹{h.groww_closing_value:,.0f} | Type: {h.holding_type}',
            'qty': h.quantity,
            'avg_price': round(h.avg_buy_price, 1),
            'current_price': round(h.groww_closing_price, 1),
            'current_value': round(h.groww_closing_value, 0),
            'pnl_pct': h.pnl_percent,
            'pnl_abs': round(h.unrealised_pnl, 0),
            'holding_type': h.holding_type,
        })

# Build MF verdicts — deduplicate by scheme name, sum values
claude_mf = fk(analysis, 'mf_verdicts') or []
mf_map = {v.get('scheme',''): v for v in claude_mf}

# Deduplicate: same scheme with multiple folios → sum values
mf_deduped = {}
for m in mfs:
    if m.scheme_name in mf_deduped:
        d = mf_deduped[m.scheme_name]
        d['units'] += m.units
        d['invested'] += m.invested_value
        d['current'] += m.current_value
        d['returns'] += m.returns_absolute
        if m.xirr > d['xirr']: d['xirr'] = m.xirr
    else:
        mf_deduped[m.scheme_name] = {
            'obj': m, 'units': m.units, 'invested': m.invested_value,
            'current': m.current_value, 'returns': m.returns_absolute, 'xirr': m.xirr,
        }

all_mf_verdicts = []
for scheme_name, data in mf_deduped.items():
    m = data['obj']
    cv = mf_map.get(scheme_name)
    rpct = (data['returns'] / data['invested'] * 100) if data['invested'] > 0 else 0
    base = {
        'scheme': scheme_name, 'amc': m.amc,
        'category': m.category, 'sub_category': m.sub_category,
        'units': round(data['units'], 2),
        'invested': round(data['invested'], 0),
        'current': round(data['current'], 0),
        'returns': round(data['returns'], 0),
        'returns_pct': round(rpct, 1),
        'xirr': data['xirr'],
    }
    if cv:
        cv.update(base)
        all_mf_verdicts.append(cv)
    else:
        verdict = 'HOLD'
        if data['xirr'] >= 12: verdict = 'CONTINUE_SIP'
        all_mf_verdicts.append({
            **base,
            'verdict': verdict,
            'action': f'XIRR: {data["xirr"]:.1f}% | Returns: ₹{data["returns"]:,.0f}',
            'reason': f'Invested: ₹{data["invested"]:,.0f} → Current: ₹{data["current"]:,.0f}',
        })

# Portfolio summary
total_inv = sum(h.buy_value for h in stocks) + sum(m.invested_value for m in mfs)
total_cur = sum(h.groww_closing_value for h in stocks) + sum(m.current_value for m in mfs)

# Build final dashboard JSON
dashboard = {
    'generated_at': analysis.get('generated_at', ''),
    'portfolio_health_score': fk(analysis, 'portfolio_health_score') or 0,
    'portfolio_summary': {
        'total_invested': round(total_inv),
        'total_current': round(total_cur),
        'total_pnl': round(total_cur - total_inv),
        'pnl_pct': round((total_cur - total_inv) / total_inv * 100, 1) if total_inv else 0,
        'stock_count': len(stocks),
        'mf_count': len(mfs),
    },
    'individual_stock_verdicts': all_stock_verdicts,
    'mf_verdicts': all_mf_verdicts,
    'top_5_urgent_actions': fk(analysis, 'top_5_urgent_actions') or [],
    'key_risks': fk(analysis, 'key_risks') or [],
}

# Market scan has the freshest recommendations — override
ms = fk(analysis, 'market_scan') or {}
if isinstance(ms, dict):
    dashboard['long_term_picks'] = ms.get('long_term_picks') or dashboard.get('long_term_picks', [])
    dashboard['dip_buying_opportunities'] = ms.get('dip_buying_opportunities', [])
    dashboard['etf_mf_recommendations'] = ms.get('etf_mf_recommendations', [])
    dashboard['aggressive_calls'] = ms.get('aggressive_calls', [])
    dashboard['intraday_setups'] = ms.get('intraday_setups') or dashboard.get('intraday_setups', [])
    dashboard['sectors_to_watch'] = ms.get('sectors_to_watch') or dashboard.get('sectors_to_watch', [])
    dashboard['promoter_signals'] = ms.get('promoter_signals') or dashboard.get('promoter_signals', [])
    dashboard['market_summary'] = ms.get('market_summary') or dashboard.get('market_summary', '')
    dashboard['fii_dii_interpretation'] = ms.get('fii_dii_interpretation') or dashboard.get('fii_dii_interpretation', '')

# Fallbacks from deep search if market_scan didn't have them
for k in ['long_term_picks','intraday_setups','sectors_to_watch','promoter_signals','market_summary','fii_dii_interpretation']:
    if not dashboard.get(k):
        v = fk(analysis, k)
        if v: dashboard[k] = v
if not dashboard.get('long_term_picks'):
    dashboard['long_term_picks'] = fk(analysis, 'new_stock_recommendations') or []
dashboard['future_multibaggers'] = fk(analysis, 'future_multibaggers') or []
dashboard['fii_dii'] = fk(analysis, 'fii_dii') or {}

# Add market indices from live_data
ld = fk(analysis, 'live_data') or {}
dashboard['market_indices'] = ld.get('indices', [])
dashboard['fii_dii'] = ld.get('fii_dii') or dashboard.get('fii_dii', {})

os.makedirs('output/reports', exist_ok=True)

# Parse order history for recent decisions analysis
recent_buys = []
active_sips = []
try:
    import openpyxl
    from collections import Counter

    # Stock order history
    if os.path.exists('input/Stocks_Order_History.xlsx'):
        wb = openpyxl.load_workbook('input/Stocks_Order_History.xlsx', data_only=True, read_only=True)
        ws = wb.active
        stock_by_name = {h.name: h for h in stocks}
        for row in ws.iter_rows(min_row=7, values_only=True):
            if not row[0] or row[0] == 'Stock name':
                continue
            try:
                name = str(row[0]).strip()
                otype = str(row[3]).strip()
                status = str(row[9]).strip() if row[9] else ''
                if otype == 'BUY' and status == 'Executed':
                    buy_price = float(str(row[5]).replace(',', '')) / float(row[4]) if row[4] and float(row[4]) > 0 else 0
                    h = stock_by_name.get(name)
                    now_price = h.groww_closing_price if h else 0
                    pnl_pct = ((now_price - buy_price) / buy_price * 100) if buy_price > 0 and now_price > 0 else 0
                    # Smart grading: compare stock drop vs Nifty drop
                    # If stock dropped similar to or less than Nifty, it's a crash victim not a bad pick
                    nifty_drop_pct = -5.0  # approximate recent Nifty correction %
                    try:
                        import json as _json
                        _idx_files = sorted([f for f in os.listdir('cache') if f.startswith('indices_')], reverse=True)
                        if _idx_files:
                            with open(os.path.join('cache', _idx_files[0])) as _f:
                                _idx = _json.load(_f)
                                for _i in _idx:
                                    if 'NIFTY 50' in str(_i.get('name', '')).upper():
                                        nifty_drop_pct = float(_i.get('change_percent', -5))
                                        break
                    except Exception:
                        pass
                    # Relative performance: how much worse than Nifty
                    relative_pct = pnl_pct - nifty_drop_pct if nifty_drop_pct < 0 else pnl_pct
                    if pnl_pct > 20:
                        grade = 'GREAT'
                    elif pnl_pct > 5:
                        grade = 'GOOD'
                    elif pnl_pct > -5:
                        grade = 'OK'
                    elif pnl_pct <= -20 and relative_pct > -10:
                        grade = 'CRASH VICTIM'  # fell with market, not fundamentally bad
                    elif pnl_pct > -20 and relative_pct > -10:
                        grade = 'MARKET DIP'  # mild drop, mostly market-driven
                    elif pnl_pct > -20:
                        grade = 'BAD'
                    else:
                        # Down >20% AND much worse than Nifty = genuinely terrible
                        grade = 'TERRIBLE' if relative_pct < -15 else 'CRASH VICTIM'
                    recent_buys.append({
                        'name': name,
                        'date': str(row[8]).strip()[:10] if row[8] else '',
                        'buy_price': round(buy_price, 1),
                        'current_price': round(now_price, 1),
                        'pnl_pct': round(pnl_pct, 1),
                        'grade': grade,
                        'value': round(float(str(row[5]).replace(',', ''))) if row[5] else 0,
                    })
            except Exception:
                pass
        wb.close()
        print(f"  ✅ Parsed {len(recent_buys)} stock buy orders")

    # MF order history
    if os.path.exists('input/MF_Order_History.xlsx'):
        wb2 = openpyxl.load_workbook('input/MF_Order_History.xlsx', data_only=True, read_only=True)
        ws2 = wb2.active
        mf_purchases = []
        for row in ws2.iter_rows(min_row=15, values_only=True):
            if not row[0] or row[0] == 'Scheme Name':
                continue
            try:
                if str(row[1]).strip() == 'PURCHASE':
                    mf_purchases.append({
                        'scheme': str(row[0]).strip(),
                        'amount': float(str(row[4]).replace(',', '')) if row[4] else 0,
                        'date': str(row[5]).strip() if row[5] else '',
                        'nav': float(row[3]) if row[3] else 0,
                    })
            except Exception:
                pass
        wb2.close()

        sip_counts = Counter(p['scheme'] for p in mf_purchases)
        # Get SIP day pattern from purchase dates
        from collections import defaultdict
        sip_days = defaultdict(list)
        for p in mf_purchases:
            try:
                dt = datetime.strptime(p['date'], '%d %b %Y')
                sip_days[p['scheme']].append(dt.day)
            except:
                pass

        for scheme, count in sip_counts.most_common(30):
            freq = 'Monthly SIP' if count >= 3 else 'Occasional'
            # Find most common SIP day
            days = sip_days.get(scheme, [])
            sip_day = Counter(days).most_common(1)[0][0] if days else 0
            active_sips.append({
                'scheme': scheme,
                'count': count,
                'frequency': freq,
                'last_amount': next((p['amount'] for p in mf_purchases if p['scheme'] == scheme), 0),
                'sip_day': sip_day,
            })
        print(f"  ✅ Detected {len([s for s in active_sips if s['frequency']=='Monthly SIP'])} active SIPs")

except Exception as e:
    print(f"  ⚠️ Order history parse error: {e}")

# Update MF verdicts with active SIP info
active_sip_names = {s['scheme'] for s in active_sips if s['frequency'] == 'Monthly SIP'}
for v in all_mf_verdicts:
    scheme = v.get('scheme', '')
    is_active_sip = scheme in active_sip_names
    xirr_val = v.get('xirr', 0) or 0

    if is_active_sip:
        v['sip_status'] = 'ACTIVE'
        if xirr_val < 5:
            v['verdict'] = 'STOP_SIP'
        elif xirr_val >= 12:
            v['verdict'] = 'CONTINUE_SIP'
        else:
            v['verdict'] = 'HOLD'
    else:
        v['sip_status'] = 'INACTIVE'
        # For inactive SIPs — suggest START if sector is good
        cat = (v.get('sub_category', '') or '') + ' ' + (v.get('category', '') or '') + ' ' + (v.get('scheme', '') or '')
        cat_upper = cat.upper()
        is_it_sector = any(kw in cat_upper for kw in ['TECHNOLOGY', 'DIGITAL INDIA', 'NASDAQ', 'US BLUECHIP', 'US TECH'])
        if is_it_sector and xirr_val < 12:
            v['verdict'] = 'START_SIP'
            v['action'] = '🔥 IT/Tech sector dip — restart SIP to average down'
        elif xirr_val >= 15:
            v['verdict'] = 'CONTINUE_SIP'
        elif xirr_val < 5:
            v['verdict'] = 'HOLD'
        else:
            v['verdict'] = 'HOLD'

# Add to dashboard
dashboard['recent_stock_buys'] = recent_buys[:30]
dashboard['active_sips'] = active_sips

# Calculate monthly SIP total
monthly_sip_total = sum(s['last_amount'] for s in active_sips if s['frequency'] == 'Monthly SIP')
dashboard['monthly_sip_total'] = monthly_sip_total

# Goal tracker: ₹1.6Cr → ₹10Cr
import math
goal = 100000000  # 10 Cr = 10,00,00,000
current = total_cur
gap = goal - current
# At 15% CAGR, how many years to reach goal?
if current > 0 and goal > current:
    years_no_sip = math.log(goal / current) / math.log(1.15)
    # With monthly SIP at 15% CAGR
    if monthly_sip_total > 0:
        # Rough: FV = PV*(1+r)^n + PMT*((1+r)^n - 1)/r
        r_monthly = 0.15 / 12
        # Binary search for years
        years_with_sip = years_no_sip
        for y in range(1, 30):
            n = y * 12
            fv = current * (1 + r_monthly) ** n + monthly_sip_total * ((1 + r_monthly) ** n - 1) / r_monthly
            if fv >= goal:
                years_with_sip = y
                break
    else:
        years_with_sip = years_no_sip
else:
    years_no_sip = 0
    years_with_sip = 0

## Acceleration scenarios — what-if analysis
def calc_years(pv, pmt_monthly, rate_annual, target):
    """How many years to reach target with lump sum + monthly SIP at given CAGR."""
    if pv >= target: return 0
    r = rate_annual / 12
    for y in range(1, 40):
        n = y * 12
        fv = pv * (1 + r) ** n + pmt_monthly * ((1 + r) ** n - 1) / r
        if fv >= target: return y
    return 40

scenarios = []
# Current path
scenarios.append({
    'label': 'Current Path',
    'sip': round(monthly_sip_total),
    'cagr': 15,
    'years': calc_years(current, monthly_sip_total, 0.15, goal),
    'extra': 0,
    'tip': 'Keep doing what you\'re doing',
})
# Increase SIP by 50%
sip_150 = monthly_sip_total * 1.5
scenarios.append({
    'label': 'Increase SIP +50%',
    'sip': round(sip_150),
    'cagr': 15,
    'years': calc_years(current, sip_150, 0.15, goal),
    'extra': round(sip_150 - monthly_sip_total),
    'tip': f'Add ₹{round((sip_150 - monthly_sip_total)/1000)}K/mo more to SIPs',
})
# Double SIP
sip_2x = monthly_sip_total * 2
scenarios.append({
    'label': 'Double SIP',
    'sip': round(sip_2x),
    'cagr': 15,
    'years': calc_years(current, sip_2x, 0.15, goal),
    'extra': round(sip_2x - monthly_sip_total),
    'tip': f'Add ₹{round((sip_2x - monthly_sip_total)/1000)}K/mo — aggressive but doable',
})
# Current SIP but 18% CAGR (better stock picking)
scenarios.append({
    'label': 'Better Returns (18%)',
    'sip': round(monthly_sip_total),
    'cagr': 18,
    'years': calc_years(current, monthly_sip_total, 0.18, goal),
    'extra': 0,
    'tip': 'Focus on quality mid-caps & multibaggers',
})
# Lump sum ₹5L + current SIP
lump_5l = 500000
scenarios.append({
    'label': 'Lump Sum ₹5L + SIP',
    'sip': round(monthly_sip_total),
    'cagr': 15,
    'years': calc_years(current + lump_5l, monthly_sip_total, 0.15, goal),
    'extra': lump_5l,
    'tip': 'Deploy ₹5L in dips (IT, quality mid-caps)',
})
# Aggressive: Double SIP + 18% CAGR
scenarios.append({
    'label': '🔥 Aggressive Mode',
    'sip': round(sip_2x),
    'cagr': 18,
    'years': calc_years(current, sip_2x, 0.18, goal),
    'extra': round(sip_2x - monthly_sip_total),
    'tip': 'Double SIP + pick multibaggers for 18% CAGR',
})

# Actionable tips based on portfolio analysis
action_tips = []
# Count EXIT/SELL stocks and their value
exit_stocks = [s for s in all_stock_verdicts if s.get('verdict','').upper() in ('EXIT','SELL')]
exit_value = sum(s.get('current_value', 0) for s in exit_stocks)
if exit_value > 0:
    action_tips.append({
        'icon': '🧹', 'title': f'Clean up {len(exit_stocks)} SELL/EXIT stocks',
        'detail': f'Free up ₹{exit_value/100000:.1f}L → redeploy into quality mid-caps or increase SIPs',
    })
# Stop SIP reallocation
stop_sips = [v for v in all_mf_verdicts if v.get('verdict','').upper() == 'STOP_SIP']
stop_sip_amt = sum(
    next((s['last_amount'] for s in active_sips if s['scheme'] == v.get('scheme','')), 0)
    for v in stop_sips
)
if stop_sip_amt > 0:
    action_tips.append({
        'icon': '🔄', 'title': f'Redirect {len(stop_sips)} underperforming SIPs',
        'detail': f'₹{stop_sip_amt/1000:.0f}K/mo from bad funds → Nifty Next 50 or Flexi Cap',
    })
# SIP step-up
action_tips.append({
    'icon': '📈', 'title': 'Annual SIP step-up 10%',
    'detail': f'Increase SIPs by 10% every year (₹{round(monthly_sip_total*0.1/1000)}K more next year)',
})
# Sector opportunities
action_tips.append({
    'icon': '💻', 'title': 'IT sector is in a dip — deploy capital',
    'detail': 'Start SIP in ICICI Tech Fund / Tata Digital India or buy TCS, Infosys on dips',
})
action_tips.append({
    'icon': '🌍', 'title': 'Add international diversification',
    'detail': 'Start SIP in Motilal Oswal NASDAQ 100 or Parag Parikh Flexi Cap (35% intl)',
})
# Consolidation
if len(stocks) > 200:
    action_tips.append({
        'icon': '🎯', 'title': f'Consolidate {len(stocks)} stocks → 50-80 quality picks',
        'detail': 'Too many positions dilute returns. Exit small positions, concentrate on winners.',
    })

dashboard['xlsx_dates'] = xlsx_dates

dashboard['goal_tracker'] = {
    'target': goal,
    'current': round(current),
    'gap': round(gap),
    'progress_pct': round(current / goal * 100, 1),
    'monthly_sip': round(monthly_sip_total),
    'yearly_sip': round(monthly_sip_total * 12),
    'years_at_15pct': round(years_no_sip, 1),
    'years_with_sip_15pct': round(years_with_sip, 1),
    'scenarios': scenarios,
    'action_tips': action_tips,
}

# Load opportunities data if available
opportunities = {}
opp_path = 'output/reports/opportunities.json'
if os.path.exists(opp_path):
    try:
        with open(opp_path) as f:
            opportunities = json.load(f)
        print(f"  ✅ Loaded {len(opportunities.get('stock_picks', []))} opportunity picks")
    except Exception:
        pass

dashboard['opportunities'] = {
    'market_regime': opportunities.get('market_regime', ''),
    'vix_assessment': opportunities.get('vix_assessment', ''),
    'sector_rotation': opportunities.get('sector_rotation', ''),
    'stock_picks': opportunities.get('stock_picks', []),
    'etf_picks': opportunities.get('etf_picks', []),
    'avoid_list': opportunities.get('avoid_list', []),
    'allocation_advice': opportunities.get('portfolio_allocation_advice', ''),
    'generated_at': opportunities.get('generated_at', ''),
}

with open('output/reports/data.json', 'w') as f:
    json.dump(dashboard, f, indent=2, ensure_ascii=False)

print(f"\n📊 Dashboard data built:")
print(f"  Stocks: {len(all_stock_verdicts)} (Claude: {len(claude_verdicts)}, auto: {len(all_stock_verdicts)-len(claude_verdicts)})")
print(f"  MFs: {len(all_mf_verdicts)}")
print(f"  Long-term picks: {len(dashboard['long_term_picks'])}")
print(f"  Intraday setups: {len(dashboard['intraday_setups'])}")
print(f"  Portfolio: ₹{total_cur:,.0f} (P&L: {dashboard['portfolio_summary']['pnl_pct']}%)")
print(f"\n💾 Saved to output/reports/data.json")
