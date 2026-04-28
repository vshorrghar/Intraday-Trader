"""Parser for Groww Stocks Holdings XLSX exports.

Reads the Groww broker stocks holdings XLSX file with header at row 11
and data rows starting from row 12. Columns B through I contain:
Stock Name, ISIN, Quantity, Average buy price, Buy value,
Closing price, Closing value, Unrealised P&L.

Each holding is classified as 'stock', 'etf', or 'invit' based on
ISIN prefix and name pattern matching.
"""

from __future__ import annotations

import logging

import openpyxl

from parsers.models import StockHolding

logger = logging.getLogger(__name__)

# Known ETF ISINs from the user's portfolio (start with INF but listed
# explicitly for clarity; the INF prefix rule already covers them).
KNOWN_ETF_ISINS: set[str] = {
    "INF457M01133",
    "INF204KB17I5",
    "INF204KB19I1",
    "INF732E01045",
    "INF204KC1402",
    "INF769K01HF4",
    "INF769K01HS7",
    "INF769K01HP3",
    "INF247L01AP3",
    "INF109KB15Y7",
    "INF666M01IO8",
    "INF179KC1HT0",
    "INF277KA1976",
    "INF277KA1984",
    "INF740KA1RE3",
}

# Known InvIT ISINs.
KNOWN_INVIT_ISINS: set[str] = {
    "INE183W23014",
    "INE0Z8Z23013",
}

# Name substrings that indicate an ETF (matched case-insensitively).
_ETF_NAME_KEYWORDS: tuple[str, ...] = (
    "ETF",
    "BEES",
    "NASDAQ",
    "MAFANG",
    "MAHKTECH",
    "SILVER",
)

# Expected header labels at row 11, columns B-I (indices 0-7 after slicing).
_EXPECTED_HEADERS: tuple[str, ...] = (
    "Stock Name",
    "ISIN",
    "Quantity",
    "Average buy price",
    "Buy value",
    "Closing price",
    "Closing value",
    "Unrealised P&L",
)

_HEADER_ROW = 11
_DATA_START_ROW = 12
_COL_START = 1  # Column A (1-indexed) — actual Groww file starts at A
_COL_END = 8    # Column H (1-indexed, inclusive)


def classify_holding(isin: str, name: str, invit_isins: set[str]) -> str:
    """Classify a holding as ``'stock'``, ``'etf'``, or ``'invit'``.

    Classification priority:
    1. ISIN present in *invit_isins* → ``'invit'``
    2. ISIN starts with ``'INF'`` → ``'etf'``
    3. Name contains any ETF keyword (case-insensitive) → ``'etf'``
    4. Everything else → ``'stock'``
    """
    if isin in invit_isins:
        return "invit"

    if isin.startswith("INF"):
        return "etf"

    name_upper = name.upper()
    for keyword in _ETF_NAME_KEYWORDS:
        if keyword in name_upper:
            return "etf"

    return "stock"


def parse_stocks_xlsx(
    file_path: str,
    invit_isins: set[str] | None = None,
) -> list[StockHolding]:
    """Parse a Groww Stocks Holdings XLSX file.

    Parameters
    ----------
    file_path:
        Path to the XLSX file exported from Groww.
    invit_isins:
        Optional set of ISINs to classify as InvIT.  Defaults to
        :data:`KNOWN_INVIT_ISINS` when *None*.

    Returns
    -------
    list[StockHolding]
        Parsed and classified holdings.

    Raises
    ------
    ValueError
        If the header row (row 11) does not match the expected structure.
    FileNotFoundError
        If *file_path* does not exist.
    """
    if invit_isins is None:
        invit_isins = KNOWN_INVIT_ISINS

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # --- Validate header at row 11 (columns B-I) -----------------------
    header_cells = [
        ws.cell(row=_HEADER_ROW, column=col).value
        for col in range(_COL_START, _COL_END + 1)
    ]
    # Normalise to strings for comparison (cells may be None).
    header_values = tuple(
        str(v).strip() if v is not None else "" for v in header_cells
    )

    if header_values != _EXPECTED_HEADERS:
        wb.close()
        raise ValueError(
            f"Unexpected header at row {_HEADER_ROW}. "
            f"Expected {_EXPECTED_HEADERS}, got {header_values}"
        )

    # --- Parse data rows from row 12 onward -----------------------------
    holdings: list[StockHolding] = []

    current_row = _DATA_START_ROW
    for row in ws.iter_rows(
        min_row=_DATA_START_ROW,
        min_col=_COL_START,
        max_col=_COL_END,
        values_only=False,
    ):
        row_num = current_row
        current_row += 1

        # Extract raw values (columns B-I → indices 0-7 in the slice).
        raw = [cell.value for cell in row]

        # Skip completely empty rows.
        if all(v is None for v in raw):
            continue

        # Validate required fields.
        try:
            name = str(raw[0]).strip()
            isin = str(raw[1]).strip()
            quantity = int(raw[2])
            avg_buy_price = float(raw[3])
            buy_value = float(raw[4])
            closing_price = float(raw[5])
            closing_value = float(raw[6])
            unrealised_pnl = float(raw[7])
        except (TypeError, ValueError) as exc:
            logger.warning(
                "Row %d: skipping due to missing/malformed value – %s", row_num, exc
            )
            continue

        if not name or not isin:
            logger.warning("Row %d: skipping due to empty name or ISIN", row_num)
            continue

        holding_type = classify_holding(isin, name, invit_isins)

        # Compute P&L percentage.
        pnl_percent = (unrealised_pnl / buy_value * 100) if buy_value else 0.0

        holdings.append(
            StockHolding(
                name=name,
                isin=isin,
                quantity=quantity,
                avg_buy_price=avg_buy_price,
                buy_value=buy_value,
                groww_closing_price=closing_price,
                groww_closing_value=closing_value,
                unrealised_pnl=unrealised_pnl,
                holding_type=holding_type,
                pnl_percent=round(pnl_percent, 2),
            )
        )

    wb.close()
    return holdings
