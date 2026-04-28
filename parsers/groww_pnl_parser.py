"""Parser for Groww P&L Report XLSX exports.

Reads the Groww broker P&L report XLSX file which contains two sheets:
- "Trade Level": individual trade records with buy dates and ISIN mappings
- "Scrip Level": aggregated scrip summaries with buy/sell quantities and P&L

The parser also computes holding periods and classifies positions as
short-term or long-term for tax purposes.
"""

from __future__ import annotations

import logging
from datetime import datetime

import openpyxl

from parsers.models import ScripSummary, TradeRecord

logger = logging.getLogger(__name__)

# Sheet names expected in the P&L XLSX.
_TRADE_SHEET = "Trade Level"
_SCRIP_SHEET = "Scrip Level"

# Tax classification thresholds in days.
_STOCK_EQUITY_MF_THRESHOLD = 365   # 12 months
_DEBT_MF_THRESHOLD = 1095          # 36 months


def compute_holding_period(buy_date: datetime, current_date: datetime) -> int:
    """Return the holding period in days between *buy_date* and *current_date*.

    Parameters
    ----------
    buy_date:
        The date the position was acquired.
    current_date:
        The reference date (typically today in IST).

    Returns
    -------
    int
        Non-negative number of days between the two dates.
    """
    delta = current_date - buy_date
    return delta.days


def classify_tax_term(holding_period_days: int, security_type: str) -> str:
    """Classify a position as ``'short_term'`` or ``'long_term'``.

    Parameters
    ----------
    holding_period_days:
        Number of days the position has been held.
    security_type:
        One of ``'stock'``, ``'equity_mf'``, or ``'debt_mf'``.

    Returns
    -------
    str
        ``'short_term'`` if below threshold, ``'long_term'`` if at or above.
    """
    if security_type == "debt_mf":
        threshold = _DEBT_MF_THRESHOLD
    else:
        # stocks and equity MF share the 12-month threshold
        threshold = _STOCK_EQUITY_MF_THRESHOLD

    return "long_term" if holding_period_days >= threshold else "short_term"


def _parse_date(value) -> datetime:
    """Parse a cell value into a datetime object.

    Handles both ``datetime`` instances (from openpyxl) and common date
    string formats.
    """
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {value!r}")


def parse_pnl_xlsx(
    file_path: str,
    current_date: datetime | None = None,
    security_type: str = "stock",
) -> tuple[list[TradeRecord], list[ScripSummary]]:
    """Parse a Groww P&L Report XLSX file.

    Parameters
    ----------
    file_path:
        Path to the P&L XLSX file exported from Groww.
    current_date:
        Reference date for holding period computation. Defaults to
        ``datetime.now()`` when *None*.
    security_type:
        Security type for tax classification (``'stock'``, ``'equity_mf'``,
        or ``'debt_mf'``). Defaults to ``'stock'``.

    Returns
    -------
    tuple[list[TradeRecord], list[ScripSummary]]
        A tuple of (trade_records, scrip_summaries).

    Raises
    ------
    ValueError
        If the expected "Trade Level" or "Scrip Level" sheet is missing.
    """
    if current_date is None:
        current_date = datetime.now()

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # Validate required sheets exist.
    missing = []
    if _TRADE_SHEET not in sheet_names:
        missing.append(_TRADE_SHEET)
    if _SCRIP_SHEET not in sheet_names:
        missing.append(_SCRIP_SHEET)
    if missing:
        wb.close()
        raise ValueError(
            f"Missing required sheet(s): {', '.join(missing)}"
        )

    trade_records = _parse_trade_sheet(wb[_TRADE_SHEET])

    # Build a map of earliest buy date per ISIN from trade records.
    buy_dates: dict[str, datetime] = {}
    for tr in trade_records:
        if tr.isin not in buy_dates or tr.trade_date < buy_dates[tr.isin]:
            buy_dates[tr.isin] = tr.trade_date

    scrip_summaries = _parse_scrip_sheet(
        wb[_SCRIP_SHEET], current_date, security_type, buy_dates
    )

    wb.close()
    return trade_records, scrip_summaries


def _parse_trade_sheet(ws) -> list[TradeRecord]:
    """Parse the 'Trade Level' sheet into TradeRecord objects.

    Expected columns (row 1 header):
    A: Stock name, B: ISIN, C: Quantity, D: Buy date, E: Buy price,
    F: Closing date, G: Closing price, H: Unrealised P&L
    """
    records: list[TradeRecord] = []
    header_row = 1

    current_row = header_row + 1
    for row in ws.iter_rows(min_row=current_row, values_only=False):
        row_num = current_row
        current_row += 1

        raw = [cell.value for cell in row]

        if all(v is None for v in raw):
            continue

        try:
            symbol = str(raw[0]).strip()
            isin = str(raw[1]).strip()
            quantity = int(raw[2])
            buy_date = _parse_date(raw[3])
            buy_price = float(raw[4])
        except (TypeError, ValueError, IndexError) as exc:
            logger.warning(
                "Trade Level row %d: skipping – %s", row_num, exc
            )
            continue

        if not symbol or not isin:
            logger.warning(
                "Trade Level row %d: skipping – empty symbol or ISIN",
                row_num,
            )
            continue

        records.append(
            TradeRecord(
                isin=isin,
                symbol=symbol,
                trade_type="buy",
                trade_date=buy_date,
                quantity=quantity,
                price=buy_price,
            )
        )

    return records


def _parse_scrip_sheet(
    ws,
    current_date: datetime,
    security_type: str,
    buy_dates: dict[str, datetime],
) -> list[ScripSummary]:
    """Parse the 'Scrip Level' sheet into ScripSummary objects.

    Expected columns (row 1 header):
    A: Stock name, B: ISIN, C: Buy quantity, D: Buy avg price,
    E: Sell quantity, F: Sell avg price, G: Realised P&L
    """
    summaries: list[ScripSummary] = []
    header_row = 1

    current_row = header_row + 1
    for row in ws.iter_rows(min_row=current_row, values_only=False):
        row_num = current_row
        current_row += 1

        raw = [cell.value for cell in row]

        if all(v is None for v in raw):
            continue

        try:
            symbol = str(raw[0]).strip()
            isin = str(raw[1]).strip()
            buy_quantity = int(raw[2])
            buy_avg_price = float(raw[3])
            sell_quantity = int(raw[4])
            sell_avg_price = float(raw[5])
            realised_pnl = float(raw[6])
        except (TypeError, ValueError, IndexError) as exc:
            logger.warning(
                "Scrip Level row %d: skipping – %s", row_num, exc
            )
            continue

        if not symbol or not isin:
            logger.warning(
                "Scrip Level row %d: skipping – empty symbol or ISIN",
                row_num,
            )
            continue

        # Look up earliest buy date from trade records.
        buy_date = buy_dates.get(isin, current_date)
        holding_days = compute_holding_period(buy_date, current_date)
        tax_class = classify_tax_term(holding_days, security_type)

        summaries.append(
            ScripSummary(
                isin=isin,
                symbol=symbol,
                buy_date=buy_date,
                buy_quantity=buy_quantity,
                buy_avg_price=buy_avg_price,
                sell_quantity=sell_quantity,
                sell_avg_price=sell_avg_price,
                realised_pnl=realised_pnl,
                holding_period_days=holding_days,
                tax_classification=tax_class,
            )
        )

    return summaries
