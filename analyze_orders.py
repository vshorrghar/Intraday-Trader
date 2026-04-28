#!/usr/bin/env python3
"""Analyze stock and MF order history for recent decisions."""
import openpyxl
from collections import Counter

# Stock orders
wb = openpyxl.load_workbook('input/Stocks_Order_History.xlsx', data_only=True, read_only=True)
ws = wb.active
stock_orders = []
for row in ws.iter_rows(min_row=7, values_only=True):
    if not row[0] or row[0] == 'Stock name':
        continue
    try:
        stock_orders.append({
            'name': str(row[0]).strip(),
            'symbol': str(row[1]).strip(),
            'type': str(row[3]).strip(),
            'qty': float(row[4]) if row[4] else 0,
            'value': float(str(row[5]).replace(',', '')) if row[5] else 0,
            'date': str(row[8]).strip() if row[8] else '',
            'status': str(row[9]).strip() if row[9] else '',
        })
    except Exception:
        pass
wb.close()

# MF orders
wb2 = openpyxl.load_workbook('input/MF_Order_History.xlsx', data_only=True, read_only=True)
ws2 = wb2.active
mf_orders = []
for row in ws2.iter_rows(min_row=15, values_only=True):
    if not row[0] or row[0] == 'Scheme Name':
        continue
    try:
        mf_orders.append({
            'scheme': str(row[0]).strip(),
            'type': str(row[1]).strip(),
            'units': float(row[2]) if row[2] else 0,
            'nav': float(row[3]) if row[3] else 0,
            'amount': float(str(row[4]).replace(',', '')) if row[4] else 0,
            'date': str(row[5]).strip() if row[5] else '',
        })
    except Exception:
        pass
wb2.close()

executed = [o for o in stock_orders if o['status'] == 'Executed']
buys = [o for o in executed if o['type'] == 'BUY']
sells = [o for o in executed if o['type'] == 'SELL']

print(f"Stock orders: {len(stock_orders)} total, {len(executed)} executed ({len(buys)} buys, {len(sells)} sells)")
print(f"MF orders: {len(mf_orders)}")

print(f"\n{'='*70}")
print("RECENT STOCK BUYS")
print(f"{'='*70}")
for o in buys[:25]:
    print(f"  {o['date'][:16]:18s} {o['name'][:32]:34s} Qty:{o['qty']:>4.0f}  Rs{o['value']:>10,.0f}")

print(f"\n{'='*70}")
print("RECENT STOCK SELLS")
print(f"{'='*70}")
for o in sells[:15]:
    print(f"  {o['date'][:16]:18s} {o['name'][:32]:34s} Qty:{o['qty']:>4.0f}  Rs{o['value']:>10,.0f}")

print(f"\n{'='*70}")
print("RECENT MF PURCHASES")
print(f"{'='*70}")
for o in mf_orders[:20]:
    print(f"  {o['date']:15s} {o['scheme'][:42]:44s} Rs{o['amount']:>8,.0f}")

# Active SIPs
sip_counts = Counter(o['scheme'] for o in mf_orders if o['type'] == 'PURCHASE')
print(f"\n{'='*70}")
print("LIKELY ACTIVE SIPs (multiple purchases = recurring SIP)")
print(f"{'='*70}")
for scheme, count in sip_counts.most_common(25):
    freq = "Monthly" if count >= 3 else "Occasional"
    print(f"  {count:>2}x  [{freq:10s}]  {scheme[:55]}")
